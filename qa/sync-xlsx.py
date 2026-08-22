"""Regenerate test-cases.xlsx from tests-data.js.

tests-data.js is the source of truth. This script parses the JS array
literal (regex, not a real JS engine) and writes a coloured Excel view.

Usage:
    python3 qa/sync-xlsx.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).resolve().parent
JS_PATH = HERE / "tests-data.js"
XLSX_PATH = HERE / "test-cases.xlsx"


def _js_object_to_json(src: str) -> str:
    """String-aware JS object literal → JSON.

    Walks the source char-by-char so bare-key detection never fires
    inside a string literal. Converts single-quoted JS strings to
    double-quoted JSON strings and quotes unquoted keys.
    """
    out: list[str] = []
    i = 0
    n = len(src)
    KEY_PRE = {"{", ",", "\n", "\t", "\r", " ", ""}

    while i < n:
        ch = src[i]

        # Double-quoted string — copy verbatim (already JSON-valid).
        if ch == '"':
            j = i + 1
            while j < n:
                if src[j] == "\\":
                    j += 2
                    continue
                if src[j] == '"':
                    j += 1
                    break
                j += 1
            out.append(src[i:j])
            i = j
            continue

        # Single-quoted JS string — convert to double-quoted.
        if ch == "'":
            j = i + 1
            body: list[str] = []
            while j < n:
                if src[j] == "\\":
                    if j + 1 < n and src[j + 1] == "'":
                        body.append("'")     # \' -> '
                        j += 2
                        continue
                    body.append(src[j])
                    body.append(src[j + 1] if j + 1 < n else "")
                    j += 2
                    continue
                if src[j] == "'":
                    j += 1
                    break
                if src[j] == '"':
                    body.append('\\"')
                else:
                    body.append(src[j])
                j += 1
            out.append('"' + "".join(body) + '"')
            i = j
            continue

        # Bare-key detection: only when preceded by [{,\n\s] (or start of stream).
        if ch.isalpha() or ch == "_":
            prev = out[-1][-1:] if out else ""
            if prev in KEY_PRE:
                j = i
                while j < n and (src[j].isalnum() or src[j] == "_"):
                    j += 1
                k = j
                while k < n and src[k] in " \t":
                    k += 1
                if k < n and src[k] == ":":
                    out.append('"' + src[i:j] + '":')
                    i = k + 1
                    continue
            out.append(ch)
            i += 1
            continue

        out.append(ch)
        i += 1

    result = "".join(out)
    result = re.sub(r",(\s*[}\]])", r"\1", result)  # trailing commas
    return result


def load_cases() -> tuple[dict, list[dict]]:
    text = JS_PATH.read_text(encoding="utf-8")
    # Strip // comments (line only)
    text = re.sub(r"^\s*//.*$", "", text, flags=re.MULTILINE)

    m_meta = re.search(r"window\.META\s*=\s*(\{.*?\});", text, re.DOTALL)
    m_cases = re.search(r"window\.TEST_CASES\s*=\s*(\[.*?\]);", text, re.DOTALL)
    if not m_meta or not m_cases:
        raise SystemExit("Could not locate META or TEST_CASES blocks in tests-data.js")

    meta = json.loads(_js_object_to_json(m_meta.group(1)))
    cases = json.loads(_js_object_to_json(m_cases.group(1)))
    return meta, cases


STATUS_FILLS = {
    "pending":  PatternFill("solid", fgColor="E5E7EB"),
    "running":  PatternFill("solid", fgColor="FEF3C7"),
    "pass":     PatternFill("solid", fgColor="DCFCE7"),
    "fail":     PatternFill("solid", fgColor="FEE2E2"),
    "blocked":  PatternFill("solid", fgColor="FFEDD5"),
    "skipped":  PatternFill("solid", fgColor="EDE9FE"),
}
STATUS_FONT_COLOR = {
    "pending": "374151",
    "running": "92400E",
    "pass":    "166534",
    "fail":    "991B1B",
    "blocked": "9A3412",
    "skipped": "5B21B6",
}

HEADER_FILL = PatternFill("solid", fgColor="1E2440")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="F3F4F6")

THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_xlsx(meta: dict, cases: list[dict]) -> None:
    wb = Workbook()

    # ------- Summary sheet -------
    ws0 = wb.active
    ws0.title = "Summary"
    ws0.column_dimensions["A"].width = 24
    ws0.column_dimensions["B"].width = 40

    ws0["A1"] = "Manu QA — Test Set"
    ws0["A1"].font = Font(size=16, bold=True)
    ws0.merge_cells("A1:B1")

    meta_rows = [
        ("Project",      meta.get("project", "")),
        ("Environment",  meta.get("environment", "")),
        ("Git SHA",      meta.get("gitSha", "")),
        ("Updated",      meta.get("updated", "")),
        ("Active layer", meta.get("activeLayer", "")),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        ws0.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws0.cell(row=i, column=2, value=v)

    # Counts
    counts = {s: 0 for s in ("pending", "running", "pass", "fail", "blocked", "skipped")}
    for c in cases:
        counts[c.get("status", "pending")] = counts.get(c.get("status", "pending"), 0) + 1
    total = len(cases)
    done = counts["pass"] + counts["fail"] + counts["blocked"] + counts["skipped"]

    ws0.cell(row=10, column=1, value="Counts").font = Font(bold=True, size=12)
    row = 11
    order = [("Total", total), ("Pending", counts["pending"]), ("Running", counts["running"]),
             ("Pass", counts["pass"]), ("Fail", counts["fail"]),
             ("Blocked", counts["blocked"]), ("Skipped", counts["skipped"]),
             ("Done", done),
             ("% Complete", f"{(done*100//total) if total else 0}%")]
    for k, v in order:
        ws0.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws0.cell(row=row, column=2, value=v)
        row += 1

    # ------- Test cases sheet -------
    ws = wb.create_sheet("Test Cases")
    headers = ["ID", "Layer", "Category", "Test Name", "Steps",
               "Expected", "Status", "Actual", "Notes"]
    widths  = [10, 8, 20, 40, 55, 55, 12, 40, 40]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BOX
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    for idx, tc in enumerate(cases, start=2):
        status = tc.get("status", "pending")
        row_data = [
            tc.get("id", ""),
            tc.get("layer", ""),
            tc.get("category", ""),
            tc.get("name", ""),
            tc.get("steps", ""),
            tc.get("expected", ""),
            status.upper(),
            tc.get("actual", ""),
            tc.get("notes", ""),
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BOX
            if col == 7:  # Status column
                cell.fill = STATUS_FILLS.get(status, STATUS_FILLS["pending"])
                cell.font = Font(bold=True, color=STATUS_FONT_COLOR.get(status, "374151"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[idx].height = 60

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(cases)+1}"

    XLSX_PATH.write_bytes(b"")  # ensure fresh
    wb.save(XLSX_PATH)


def main() -> None:
    meta, cases = load_cases()
    write_xlsx(meta, cases)
    total = len(cases)
    done = sum(1 for c in cases if c.get("status") in ("pass", "fail", "blocked", "skipped"))
    print(f"Wrote {XLSX_PATH.name}: {total} cases, {done} done, {total-done} pending/running")


if __name__ == "__main__":
    main()
